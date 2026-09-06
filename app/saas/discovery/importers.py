"""In-memory CSV/XLSX importers that emit tenant-neutral job dictionaries."""

from __future__ import annotations

import csv
import io
import re
import zipfile
from pathlib import PurePath
from typing import Any
from urllib.parse import urlsplit

from .common import (
    NormalizedJob,
    clean_text,
    humanize_slug,
    make_job,
    safe_http_url,
    stable_external_id,
)
from .providers import detect_provider


MAX_IMPORT_BYTES = 6 * 1024 * 1024
MAX_IMPORT_ROWS = 2_000
MAX_IMPORT_COLUMNS = 200
MAX_XLSX_EXPANDED_BYTES = 40 * 1024 * 1024


HEADER_SYNONYMS: dict[str, frozenset[str]] = {
    "company": frozenset({
        "company", "company name", "company_name", "organization", "organisation", "org", "employer"
    }),
    "title": frozenset({
        "role", "title", "position", "job title", "job_title", "job", "role title", "opening"
    }),
    "email": frozenset({
        "email", "email address", "email_address", "e-mail", "contact email", "recruiter email", "mail"
    }),
    "location": frozenset({"location", "city", "place", "loc", "work location"}),
    "description": frozenset({
        "description", "job description", "job_description", "jd", "details", "requirements"
    }),
    "apply_url": frozenset({
        "apply_url", "apply url", "apply link", "application url", "application link",
        "job url", "job_url", "link", "url", "apply"
    }),
    "external_id": frozenset({"external id", "external_id", "job id", "job_id", "requisition id", "req id"}),
    "source": frozenset({"source", "job source", "board", "platform"}),
    "domain": frozenset({"domain", "website", "site", "company website", "url domain"}),
    "company_url": frozenset({"company url", "company_url", "company website url", "employer website"}),
    "salary": frozenset({"salary", "ctc", "compensation", "pay", "package", "stipend", "lpa"}),
    "verified": frozenset({"verified", "is verified", "trusted", "confirmed"}),
    # External-AI research workbooks use these fields to carry contact
    # provenance and the exact JD into the review-gated outreach flow.
    "contact_name": frozenset({
        "contact name", "contact_name", "person name", "person_name", "recruiter name",
        "recruiter_name", "hiring manager", "hiring_manager", "recipient name",
    }),
    "contact_title": frozenset({
        "contact title", "contact_title", "person title", "person_title", "recruiter title",
        "recruiter_title", "job title at company", "contact role",
    }),
    "contact_type": frozenset({
        "contact type", "contact_type", "recipient type", "recipient_type",
    }),
    "email_verification_status": frozenset({
        "email verification status", "email_verification_status", "verification status",
        "verification_status", "email status", "email_status",
    }),
    "linkedin_url": frozenset({
        "linkedin", "linkedin url", "linkedin_url", "person linkedin", "person_linkedin",
    }),
    "source_url": frozenset({"source url", "source_url", "evidence url", "evidence_url"}),
    "contact_source_url": frozenset({
        "contact source url", "contact_source_url", "email source url", "email_source_url",
        "contact evidence url", "contact_evidence_url",
    }),
    "source_date": frozenset({"source date", "source_date", "evidence date", "evidence_date"}),
    "contact_source": frozenset({"contact source", "contact_source", "email source", "email_source"}),
    "experience_required": frozenset({
        "experience required", "experience_required", "years required", "years_required",
        "required experience", "required_experience", "minimum experience", "minimum_experience",
    }),
}


def _norm_header(value: object) -> str:
    text = clean_text(value).lower().replace("_", " ")
    return re.sub(r"\s+", " ", text)


def _header_map(row: list[Any]) -> dict[str, int]:
    mapping: dict[str, int] = {}
    for index, value in enumerate(row):
        normalized = _norm_header(value)
        for canonical, aliases in HEADER_SYNONYMS.items():
            if normalized in aliases and canonical not in mapping:
                mapping[canonical] = index
    return mapping


def _value(row: list[Any], mapping: dict[str, int], key: str) -> str:
    index = mapping.get(key)
    if index is None or index >= len(row):
        return ""
    return clean_text(row[index], limit=25_000)


def _derive_company(domain: str, email: str, apply_url: str | None) -> str:
    candidate = domain.strip().lower()
    if not candidate and "@" in email:
        candidate = email.rsplit("@", 1)[1]
    if not candidate and apply_url:
        parsed = urlsplit(apply_url)
        segments = [segment for segment in parsed.path.split("/") if segment]
        if detect_provider(apply_url) in {"greenhouse", "lever", "ashby"} and segments:
            candidate = segments[0]
        else:
            candidate = (parsed.hostname or "").split(".")[0]
    candidate = candidate.split(".")[0]
    return humanize_slug(candidate) or "Unknown employer"


def _matrix_to_jobs(matrix: list[list[Any]], *, max_rows: int) -> list[NormalizedJob]:
    header_index: int | None = None
    mapping: dict[str, int] = {}
    for index, row in enumerate(matrix[:50]):
        candidate = _header_map(row)
        if candidate:
            header_index, mapping = index, candidate
            break
    if header_index is None:
        return [] if not any(any(clean_text(cell) for cell in row) for row in matrix) else _raise_headers()

    jobs: list[NormalizedJob] = []
    seen: set[str] = set()
    for row_number, row in enumerate(matrix[header_index + 1 :], start=header_index + 2):
        if len(jobs) >= max_rows:
            break
        if not any(clean_text(cell) for cell in row):
            continue
        title = _value(row, mapping, "title") or "Open position"
        email = _value(row, mapping, "email")
        apply_url = safe_http_url(_value(row, mapping, "apply_url"))
        domain = _value(row, mapping, "domain")
        company = _value(row, mapping, "company") or _derive_company(domain, email, apply_url)
        location = _value(row, mapping, "location") or None
        description = _value(row, mapping, "description")
        salary = _value(row, mapping, "salary")
        raw_external_id = _value(row, mapping, "external_id")
        source = _value(row, mapping, "source") or "file_import"
        verified_value = _norm_header(_value(row, mapping, "verified"))
        email_status = _norm_header(_value(row, mapping, "email_verification_status"))
        if not description:
            details = [f"{title} opportunity at {company}."]
            if location:
                details.append(f"Location: {location}.")
            if salary:
                details.append(f"Compensation: {salary}.")
            description = " ".join(details)
        external_id = raw_external_id or stable_external_id(
            "file_import", apply_url, email.lower(), company.lower(), title.lower()
        )
        if external_id in seen:
            continue
        seen.add(external_id)
        provider = detect_provider(apply_url) if apply_url else None
        jobs.append(
            make_job(
                source=source,
                external_id=external_id,
                apply_url=apply_url,
                title=title,
                company=company,
                location=location,
                description=description,
                contact_email=email,
                metadata={
                    "provider": provider,
                    "import_row": row_number,
                    "domain": domain,
                    "company_url": safe_http_url(_value(row, mapping, "company_url")),
                    "compensation": salary,
                    "verified": verified_value
                    in {"1", "true", "yes", "y", "verified", "trusted"}
                    or email_status
                    in {
                        "public source verified",
                        "source verified",
                        "publicly listed",
                        "public source",
                    },
                    "contact_name": _value(row, mapping, "contact_name")[:160] or None,
                    "contact_title": _value(row, mapping, "contact_title")[:160] or None,
                    "contact_type": _value(row, mapping, "contact_type")[:80] or None,
                    "email_verification_status": (
                        _value(row, mapping, "email_verification_status")[:80] or None
                    ),
                    "linkedin_url": safe_http_url(_value(row, mapping, "linkedin_url")),
                    "source_url": safe_http_url(_value(row, mapping, "source_url")),
                    "contact_source_url": safe_http_url(
                        _value(row, mapping, "contact_source_url")
                    ),
                    "source_date": _value(row, mapping, "source_date")[:40] or None,
                    "contact_source": _value(row, mapping, "contact_source")[:240] or None,
                    "experience_required": _value(row, mapping, "experience_required")[:160] or None,
                    "external_research": True,
                },
            )
        )
    return jobs


def _raise_headers() -> list[NormalizedJob]:
    raise ValueError("No recognized job columns were found")


def _bounded_data(data: object) -> bytes:
    if not isinstance(data, bytes):
        raise TypeError("Uploaded spreadsheet must be bytes")
    if not data:
        return b""
    if len(data) > MAX_IMPORT_BYTES:
        raise ValueError("Uploaded spreadsheet is too large")
    return data


def parse_csv_bytes(data: bytes, *, max_rows: int = MAX_IMPORT_ROWS) -> list[NormalizedJob]:
    """Parse a UTF-8 CSV with case-insensitive, flexible column names."""

    payload = _bounded_data(data)
    if not payload:
        return []
    text = payload.decode("utf-8-sig", "replace")
    sample = text[:16_384]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
    except csv.Error:
        dialect = csv.excel
    bounded_rows = max(1, min(int(max_rows), MAX_IMPORT_ROWS))
    try:
        matrix: list[list[str]] = []
        for row_number, row in enumerate(csv.reader(io.StringIO(text), dialect=dialect)):
            if row_number >= bounded_rows + 50:
                break
            matrix.append(row)
    except (csv.Error, UnicodeError) as exc:
        raise ValueError("CSV could not be parsed") from exc
    return _matrix_to_jobs(matrix, max_rows=bounded_rows)


def _validate_xlsx_archive(payload: bytes) -> None:
    try:
        with zipfile.ZipFile(io.BytesIO(payload)) as archive:
            total = 0
            for info in archive.infolist():
                total += info.file_size
                if total > MAX_XLSX_EXPANDED_BYTES:
                    raise ValueError("Expanded XLSX workbook is too large")
                if info.file_size > 5_000_000 and info.compress_size * 200 < info.file_size:
                    raise ValueError("XLSX workbook has an unsafe compression ratio")
    except zipfile.BadZipFile as exc:
        raise ValueError("XLSX workbook is not a valid Office archive") from exc


def parse_xlsx_bytes(data: bytes, *, max_rows: int = MAX_IMPORT_ROWS) -> list[NormalizedJob]:
    """Parse the first recognizable worksheet without formulas or filesystem writes.

    External AI tools often put an instructions/README sheet before the data
    sheet.  Selecting by recognized headers keeps that harmless sheet from
    making an otherwise valid workbook look empty or malformed.
    """

    payload = _bounded_data(data)
    if not payload:
        return []
    _validate_xlsx_archive(payload)
    try:
        import openpyxl
    except ImportError as exc:  # pragma: no cover - deployment dependency guard
        raise RuntimeError("XLSX import requires openpyxl") from exc
    try:
        workbook = openpyxl.load_workbook(
            io.BytesIO(payload), read_only=True, data_only=True, keep_links=False
        )
    except Exception as exc:  # noqa: BLE001 - openpyxl exposes several format errors
        raise ValueError("XLSX workbook could not be parsed") from exc
    try:
        bounded_rows = max(1, min(int(max_rows), MAX_IMPORT_ROWS))
        matrix: list[list[Any]] | None = None
        worksheets = list(workbook.worksheets)[:10]
        for worksheet in worksheets:
            if (worksheet.max_column or 1) > MAX_IMPORT_COLUMNS:
                raise ValueError("XLSX worksheet has too many columns")
            candidate = [
                list(row)
                for row in worksheet.iter_rows(
                    min_row=1,
                    max_row=min(worksheet.max_row or 1, bounded_rows + 50),
                    max_col=min(worksheet.max_column or 1, MAX_IMPORT_COLUMNS),
                    values_only=True,
                )
            ]
            if any(_header_map(row) for row in candidate[:50]):
                matrix = candidate
                break
    finally:
        workbook.close()
    if matrix is None:
        raise ValueError("No worksheet with recognized job columns was found")
    return _matrix_to_jobs(matrix, max_rows=bounded_rows)


def parse_spreadsheet_bytes(
    data: bytes, filename: str, *, max_rows: int = MAX_IMPORT_ROWS
) -> list[NormalizedJob]:
    """Dispatch an uploaded `.csv` or `.xlsx` file by its final suffix."""

    suffix = PurePath(filename or "").suffix.lower()
    if suffix == ".csv":
        return parse_csv_bytes(data, max_rows=max_rows)
    if suffix == ".xlsx":
        return parse_xlsx_bytes(data, max_rows=max_rows)
    raise ValueError("Only .csv and .xlsx job imports are supported")


__all__ = [
    "HEADER_SYNONYMS",
    "MAX_IMPORT_BYTES",
    "MAX_IMPORT_COLUMNS",
    "MAX_IMPORT_ROWS",
    "parse_csv_bytes",
    "parse_spreadsheet_bytes",
    "parse_xlsx_bytes",
]
